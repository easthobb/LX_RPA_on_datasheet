import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook


## iteration 변수
iter = 4

## 건물 양식 입력 자동화 python file

# 엑셀파일 열기
data = openpyxl.load_workbook('desktop/rpa/data.xlsx') #로 데이터 시트 오픈
form = openpyxl.load_workbook('desktop/rpa/RPA_form.xlsx') #폼 데이터 시트 오픈

# 로 데이터 시트와 폼 시트의 fopen
data_sheet = data.get_sheet_by_name('building') #로 데이터 시트 객체 파일(building,시트명 : building)
form_sheet_page_1 = form.get_sheet_by_name('building_1') #폼 시트 페이지 1(building, 시트명 : building_1)
form_sheet_page_2 = form.get_sheet_by_name('building_2') #폼 시트 페이지 2(building, 시트명 : building_2)

# (예정) for 문 삽입부분
# 로 데이터 개별 객체를 데이터 시트에서 추출
for i in range(iter):
    sheet_start_cell = "A" + str(i+iter)
    sheet_end_cell = "AJ" + str(i+iter)
    
    building_object = data_sheet[sheet_start_cell:sheet_end_cell] # 반복시 변경해야 됩니다
    data_list = []
    for row in building_object:
        for cell in row:
            data_list.append(cell.value)

    print(data_list)
# (예정) for 문 삽입부분
# 추출한 데이터를 양식에 매핑시켜주는 부분
    form_sheet_page_1['B2'] = data_list[0] ##관리번호
    form_sheet_page_1['D6'] = data_list[1] ##고유번호
    form_sheet_page_1['J6'] = data_list[2] ## 재산번호
    form_sheet_page_1['D7'] = data_list[3] ## 소재지
    form_sheet_page_1['J7'] = data_list[4] ## 재산명칭
    form_sheet_page_1['D8'] = data_list[5] ## 공유자수
    form_sheet_page_1['G8'] = data_list[6] ## 공유지분
    form_sheet_page_1['J8'] = data_list[7] ## 관리관
    form_sheet_page_1['D9'] = data_list[8] ## 재산구분
    form_sheet_page_1['J9'] = data_list[9] ## 위임기관

    # (예정) for문 삽입부분 
    # 이미지 입력부분 : 리사이즈
    for j in range(4):
        img_file_name = 'desktop/rpa/image/' + str(data_list[0]) +'-' +str(j+1) +'.png' # 첫번째 페이지 이미지 이름 지정
        print(img_file_name)
        img = openpyxl.drawing.image.Image(img_file_name)
        img.width=430 # 이미지 리사이징, 가로.픽셀 단위입니다.
        img.height=286 # 이미지 리사이징 세로.픽셀 단위입니다.
        if(j==0):
            form_sheet_page_1.add_image(img,'B20')
        elif(j==1):
            form_sheet_page_1.add_image(img,'H20')
        elif(j==2):
            form_sheet_page_1.add_image(img,'B34')
        elif(j==3):
            form_sheet_page_1.add_image(img,'H34')

    output_file_name = 'desktop/rpa/output/output' + str(data_list[0]) +'.xlsx'
    form.save(output_file_name)

    # sample_img.anchor(form_work_sheet.cell('B6'))
    # form_work_sheet.add_image(sample_img,'B6')

#form.save('desktop/rpa/output/output.xlsx')

form.close()
data.close()
#work_sheet_new = work_book.create_sheet('new sheet')

# work_sheet.rows는 해당 쉬트의 모든 행을 객체로 가지고 있음
# for each_row in work_sheet.rows:
#     # cell(row=행 번호, column=열 번호).value = 해당 세로셀/가로셀에 어떤 값을 넣어주세요
#     work_sheet_new.cell(row=each_row[0].row, column=1).value = each_row[2].value
    
# 엑셀 파일 저장
# work_book.save("desktop/rpa/output.xlsx")
# work_book.close()