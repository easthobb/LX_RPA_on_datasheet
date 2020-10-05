import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Alignment
import win32com.client
import time

## 토지 양식 입력 자동화 python file

###출력파일 및 excel PDF 파일 절대경로 설정
filepath ="C:\\Users\\user\\Desktop\\RPA\\"
input_image_format = '.png'
saved_name = '실태조사_'

### data sheet에서 필요한 설정요소 ### 
data = openpyxl.load_workbook(filepath+'data.xlsx') #로 데이터 시트 오픈(변경가능)
data_sheet_name = 'land' ## 데이터 파일 시트 
start_column = 'A' ## 데이터 파일 시트에서 값을 가져올 시작 열
end_column = 'AP' ## 데이터 파일 시트에서 값을 가져올 끝 열
iter = 0 ## 문서 생성을 할 데이터의 갯수 
data_sheet = data.get_sheet_by_name(data_sheet_name) #로 데이터 시트 객체 파일(land,시트명 : land)

for row in data_sheet.rows:
    iter=iter+1

## 엑셀 PDF 연결
excel = win32com.client.Dispatch('Excel.Application')
excel.Visible = False
excel.ScreenUpdating = False
excel.DisplayAlerts = False
excel.EnableEvents = False

# 로 데이터 개별 객체를 데이터 시트에서 추출
for i in range(iter): 
    
    ### form sheet 에서 필요한 설정요소
    form = openpyxl.load_workbook(filepath+'RPA_form.xlsx') #폼 데이터 시트 경로 지정(변경가능)
    form_sheet_list = ['land_1','land_2','land_3'] ## 접근할 시트(추가 및 삭제 가능)
    form_sheet_page_1 = form.get_sheet_by_name(form_sheet_list[0]) #폼 시트 페이지 1(building, 시트명 : land_1)
    form_sheet_page_2 = form.get_sheet_by_name(form_sheet_list[1]) #폼 시트 페이지 2(building, 시트명 : land_2)
    form_sheet_page_3 = form.get_sheet_by_name(form_sheet_list[2]) #폼 시트 페이지 3(building, 시트명 : land_3)

    sheet_start_cell = start_column + str(i+4) 
    sheet_end_cell = end_column + str(i+4)
    
    land_object = data_sheet[sheet_start_cell:sheet_end_cell] # 반복시 변경해야 됩니다
    data_list = [] ## 개별 토지 객체의 정보가 담기는 리스트 생성
    for row in land_object: 
        for cell in row:
            data_list.append(cell.value)

    print(data_list) # 각 행의 값들을 콘솔에 표시
    key_value = str(data_list[0]) #넘버링을 위한 key value 엑셀에 첫번째 행에 해당합니다. , 예상 정수값
    
    ## 수임번호가 없을 경우 해당 셀 pass
    if(data_list[10]==None):
        print('조사연번'+str(data_list[0])+'에 해당하는 행의 수임번호가 없습니다.')
        continue

    # (예정) for 문 삽입부분
    # 추출한 데이터를 양식에 매핑시켜주는 부분 :  land_1 sheet
    form_sheet_page_1['B1'] = data_list[0] ## 관리번호
    form_sheet_page_1['D6'] = data_list[1] ## 고유번호
    form_sheet_page_1['J6'] = data_list[2] ## 재산번호
    form_sheet_page_1['D7'] = data_list[3] ## 소재지
    form_sheet_page_1['J7'] = data_list[4] ## 재산명칭
    form_sheet_page_1['D8'] = data_list[5] ## 공유자수
    form_sheet_page_1['G8'] = data_list[6] ## 공유지분
    form_sheet_page_1['J8'] = data_list[7] ## 관리관
    form_sheet_page_1['D9'] = data_list[8] ## 재산구분
    form_sheet_page_1['J9'] = data_list[9] ## 위임기관

    form_sheet_page_1['D10'] = data_list[10] ## 대장면적
    form_sheet_page_1['J10'] = data_list[11] ## 대장지목
    form_sheet_page_1['D11'] = data_list[12] ## 이용면적
    form_sheet_page_1['J11'] = data_list[13] ## 이용지목
    form_sheet_page_1['D12'] = data_list[14] ## 공시지가
    form_sheet_page_1['J12'] = data_list[15] ## 점유형태
    form_sheet_page_1['D13'] = data_list[16] ## 취득일자
    form_sheet_page_1['J13'] = data_list[17] ## 점유필지
    form_sheet_page_1['D14'] = data_list[18] ## 사용형태
    form_sheet_page_1['D15'] = data_list[19] ## 조사현황
    form_sheet_page_1['C16'] = data_list[20] ## 토지이용계획
  
    ## 소재지 00시 00구 추출 부분 
    form_sheet_page_1['C1'] = " " + data_list[3] + ' 단독사용[토지]'
        
    #추출한 데이터를 매핑시켜주는 부분 :  building_2 sheet

    form_sheet_page_2['B16'] = data_list[21] ## 기호
    form_sheet_page_2['C16'] = data_list[22] ## 사용형태(계약방법)
    form_sheet_page_2['D16'] = data_list[23] ## 이용상황
    form_sheet_page_2['F16'] = data_list[24] ## 점유면적
    form_sheet_page_2['H16'] = data_list[25] ## 계약일
    form_sheet_page_2['J16'] = data_list[26] ## 계약기간
    form_sheet_page_2['L16'] = data_list[27] ## 대부료
    form_sheet_page_2['M16'] = data_list[28] ## 비고
    form_sheet_page_2['D22'] = data_list[29] ## 접근성
    form_sheet_page_2['D23'] = data_list[30] ## 주요시설

    form_sheet_page_2['D27'] = data_list[31] ## 1차분류=임시
    form_sheet_page_2['F27'] = data_list[32] ## 1차분류=임시
    form_sheet_page_2['H27'] = data_list[33] ## 1차분류=임시
    form_sheet_page_2['J27'] = data_list[34] ## 1차분류=임시
    form_sheet_page_2['L27'] = data_list[35] ## 1차분류=임시
    form_sheet_page_2['D28'] = data_list[36] ## 2차분류=임시
    form_sheet_page_2['F28'] = data_list[37] ## 2차분류=임시
    form_sheet_page_2['H28'] = data_list[38] ## 2차분류=임시
    form_sheet_page_2['J28'] = data_list[39] ## 2차분류=임시
    form_sheet_page_2['L28'] = data_list[40] ## 2차분류=임시
    form_sheet_page_2['D29'] = data_list[41] ## 활용의견
    
    # 이미지 입력부분 : 리사이즈 , 이미지 파일명 : 관리번호-1( ),관리번호-2( ), 관리번호-3( ),관리번호-4( ), 관리번호-5( ),
    # (향후 변경) 1 : 지적도 , 2 : 국토정보기본도, 3 : 현황사진, 4 :사용허가및무단점유현황, 5 : 토지이용계획확인서 
    for j in range(5):
        img_file_name = filepath + '\\image\\' +  key_value +'_' +str(j+1) + input_image_format # 첫번째 페이지 이미지 이름 지정
        print(img_file_name)
        try: ##이미지 파일 삽입시도.
            img = openpyxl.drawing.image.Image(img_file_name)
        except: ##해당하는 이미지 파일이 없을경우, form overriding 방지
            print('수임번호 '+ key_value +'에 해당하는 이미지 파일이 존재하지 않습니다.')
            if(j==0): # 지적도
                print('지적도 누락!')
            elif(j==1): # 국토정보기본도
                print('국토정보기본도 누락!')
            elif(j==2): # 현황사진
                print('현황사진 누락!')
            elif(j==3):#사용허가및무단점유현황
                print('사용허가 및 무단점유현황 누락')
            elif(j==4):# 토지이용계획확인서
                print('토지이용계획확인서 누락')
        else:
            if(j==0): # 지적도
                img.width=430 # 이미지 리사이징, 가로.픽셀 단위입니다.
                img.height=286 # 이미지 리사이징 세로.픽셀 단위입니다.
                form_sheet_page_1.add_image(img,'B20')
            elif(j==1): # 국토정보기본도
                img.width=430 # 이미지 리사이징, 가로.픽셀 단위입니다.
                img.height=286 # 이미지 리사이징 세로.픽셀 단위입니다.
                form_sheet_page_1.add_image(img,'H20')
            elif(j==2): # 현황사진
                img.width=430 # 이미지 리사이징, 가로.픽셀 단위입니다.
                img.height=286 # 이미지 리사이징 세로.픽셀 단위입니다.
                form_sheet_page_1.add_image(img,'E34')
            elif(j==3): # 사용허가및무단점유현황
                img.width=864 # 이미지 리사이징, 가로.픽셀 단위입니다.
                img.height=400 # 이미지 리사이징 세로.픽셀 단위입니다.
                form_sheet_page_2.add_image(img,'B3')
            elif(j==4):
                img.width=864 # 이미지 리사이징, 가로.픽셀 단위입니다.
                img.height=816 # 이미지 리사이징 세로.픽셀 단위입니다.
                form_sheet_page_3.add_image(img,'B3')

    #output file 저장 부분
    output_file_name = filepath + '\\output\\' + saved_name + key_value + '(토지)' + '.xlsx' ## data_list[0] 관리번호
    form.save(output_file_name)

    #output file PDF 저장
    wb = excel.WorkBooks.Open(output_file_name)
    ws_chart = wb.WorkSheets(form_sheet_list)
    ws_chart.Select()
    pdf_save_path = filepath + '\\output\\' + saved_name + key_value + '(토지)' + '.pdf'
    wb.ActiveSheet.ExportAsFixedFormat(0,pdf_save_path)
    wb.Close(True)

form.close()
data.close()
excel.Quit()
print('모든 변환이 완료되었습니다. \n 엔터를 누르시면 프로그램을 종료합니다.')
a=input()