import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Alignment
import win32com.client
import PIL



## 토지 양식 입력 자동화 python file
## version 1.4 (2020/11/10) 
###출력파일 및 excel PDF 파일 절대경로 설정
filepath ="D:\workingset\\"
input_image_format = '.jpg'
saved_name = '실태조사_'


## result DB 조건
## 날짜 데이터 포맷
## 수임번호 줄바꿈 없어야함


# 사용자에게 시트 입력을 받는 부분
while(1):
    print('대상 재산이 시내 시트에 존재하는 데이터일 경우 1 을 입력해주시고 \n 시계외 데이터일 경우 0 을 입력해주세요!(enter)')
    temp = input()
    if(temp =='1'):
        data_sheet_name = '시내'
        break
    elif(temp=='0'):
        data_sheet_name = '시계외'
        break
    else:
        print('잘못입력하셨습니다. 다시 입력해주세요 :)')


### data sheet에서 필요한 설정요소 ### 
data = openpyxl.load_workbook(filepath+'resultDB.xlsx') #로 데이터 시트 오픈(변경가능)
start_column = 'A' ## 데이터 파일 시트에서 값을 가져올 시작 열
end_column = 'R' ## 데이터 파일 시트에서 값을 가져올 끝 열
iter = 0 ## 문서 생성을 할 데이터의 갯수
data_sheet = data.get_sheet_by_name(data_sheet_name) #로 데이터 파일(resultDB.xlxs,시트명 : '시계외' or '시내')

for row in data_sheet.rows:
    iter=iter+1

## 엑셀 PDF 연결, background 실행
excel = win32com.client.Dispatch('Excel.Application')
excel.Visible = False
excel.ScreenUpdating = False
excel.DisplayAlerts = False
excel.EnableEvents = False
excel.Interactive = False
# 로 데이터 개별 객체를 데이터 시트에서 추출
for i in range(iter):

    ### form sheet 에서 필요한 설정요소
    form = openpyxl.load_workbook(filepath+'RPA_form.xlsx') #폼 데이터 시트 경로 지정(변경가능)
    form_sheet_list = ['simple'] ## 접근할 시트(추가 및 삭제 가능)
    form_sheet_page_1 = form.get_sheet_by_name(form_sheet_list[0]) #폼 시트 페이지 1(simple, 시트명 : simple)
    
    ### 시작셀과 마지막 셀 설정 
    sheet_start_cell = start_column + str(i+4) ## 하드코딩 시작 데이터 셀 ex) A4 -> i = 4 , A7 -> i = 7
    sheet_end_cell = end_column + str(i+4)
    
    land_object = data_sheet[sheet_start_cell:sheet_end_cell] # 반복시 변경해야 됩니다
    data_list = [] ## 개별 토지 객체의 정보가 담기는 리스트 생성
    for row in land_object: 
        for cell in row:
            data_list.append(cell.value)

    print('가져오는 중 : ' + str(data_list)) # 각 행의 값들을 콘솔에 표시
    key_value = str(data_list[10]) #넘버링을 위한 key value 엑셀에 ****수임번호****에 해당합니다. , 예상 정수값
    



    ## 수임번호가 없을 경우 해당 셀 pass
    if(data_list[10]==None):
        print('조사연번'+str(data_list[0])+'에 해당하는 행의 수임번호가 없습니다.\n 생성 실패.')
        continue
    
    # (예정) for 문 삽입부분
    # 추출한 데이터를 양식에 매핑시켜주는 부분 :  land_1 sheet
    # form_sheet_page_1['B7'] = data_list[0] ## 조사연번
    form_sheet_page_1['C2'] = "   "+ data_list[1] ## 소재지
    form_sheet_page_1['D7'] = data_list[2] ## 대장지목
    form_sheet_page_1['J7'] = str(data_list[3]) + '㎡'## 대장면적
    form_sheet_page_1['D8'] = str(format(int(data_list[4]),",")) + '원/㎡'## 공시지가(20년)
    form_sheet_page_1['J8'] = str(data_list[5]).split(' ')[0] ## 취득일자
    #form_sheet_page_1['D9'] = data_list[6] ## 공유지분여부 - form에 없음
    form_sheet_page_1['D9'] = data_list[7] ## 공유자수(공유자수 시 포함)
    form_sheet_page_1['J9'] = data_list[8] ## 서울시 공유지분
    form_sheet_page_1['D10'] = data_list[9] ## PNU
    form_sheet_page_1['B2'] = data_list[10] ## 수임번호 - 라벨
    form_sheet_page_1['J10'] = data_list[10] ## 수임번호 - 시트
    form_sheet_page_1['D11'] = str(data_list[11]).split(' ')[0] ## 수탁일자
    form_sheet_page_1['J11'] = data_list[12] ## 재산관리관
    form_sheet_page_1['D12'] = data_list[13] ## 사용형태
    form_sheet_page_1['D13'] = str(data_list[14]).split(' ')[0] ## 조사일자
    form_sheet_page_1['D14'] = data_list[15] ## 조사내용
    form_sheet_page_1['D15'] = data_list[16] ## 토지이용계획 - form에 없음
    form_sheet_page_1['J13'] = data_list[17] ## 현황측량실시여부
    
    #1104 이미지 파일 매핑 때문에 넣어줌
    img_name_form = ' '.join(data_list[1].split()[-2:])
    
    # 이미지 입력부분 : 리사이즈 , 이미지 파일명 : 수임번호_1( ), 수임번호-2( ), 수임번호_3( )
    # (향후 변경) 1 : 지적도 , 2 : 국토정보기본도, 3 : 현황사진
    for j in range(3):
        img_file_name = filepath + '\\image\\' + img_name_form + '_' +str(j+1) + input_image_format # 이미지경로 
        print('이미지 파일 삽입중 : ' + img_file_name)
        try: ##이미지 파일 삽입시도.
            img = openpyxl.drawing.image.Image(img_file_name)
        except: ##해당하는 이미지 파일이 없을경우, form overriding 방지
            print('수임번호'+ key_value +'에 해당하는 이미지 파일이 존재하지 않습니다.')
            if(j==0): # 지적도
                print('지적도 누락!')
            elif(j==1): # 국토정보기본도
                print('국토정보기본도 누락!')
            elif(j==2): # 현황사진
                print('현황사진 누락!')
        else:
            if(j==0): # 지적도
                img.width=473 # 이미지 리사이징, 가로.픽셀 단위입니다.
                img.height=290 # 이미지 리사이징 세로.픽셀 단위입니다.
                form_sheet_page_1.add_image(img,'B20') ## 이미지가 들어갈 셀
            elif(j==1): # 국토정보기본도
                img.width=473    # 이미지 리사이징, 가로.픽셀 단위입니다.
                img.height=290 # 이미지 리사이징 세로.픽셀 단위입니다.
                form_sheet_page_1.add_image(img,'H20') ## 이미지가 들어갈 셀
            elif(j==2): # 현황사진
                img.width=442 # 이미지 리사이징, 가로.픽셀 단위입니다.
                img.height=290 # 이미지 리사이징 세로.픽셀 단위입니다.
                form_sheet_page_1.add_image(img,'E34') ## 이미지가 들어갈 셀
        

    #output file 저장 부분
    output_file_name = filepath + '\\output\\' + saved_name + key_value +'.xlsx' ## data_list[10] 수임번호
    form.save(output_file_name)
    form.close()

    #output file PDF 저장
    excel.Visible = False
    wb = excel.WorkBooks.Open(output_file_name)
    excel.Visible = False
    ws_chart = wb.WorkSheets(form_sheet_list)
    ws_chart.Select()
    pdf_save_path = filepath + "\\output\\" +saved_name + key_value + '.pdf'
    wb.ActiveSheet.ExportAsFixedFormat(0,pdf_save_path)
    wb.Close(True)

data.close()
excel.Quit()
print('모든 변환이 완료되었습니다. \n 엔터를 누르시면 프로그램을 종료합니다.')
a=input()