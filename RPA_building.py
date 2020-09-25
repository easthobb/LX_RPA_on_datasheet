import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Alignment
import win32com.client
import time

## 건물 양식 입력 자동화 python file

###출력파일 및 excel PDF 파일 절대경로 설정
filepath ="C:\\Users\\user\\Desktop\\RPA\\"
input_image_format = '.png'
saved_name = 'building'

### data sheet에서 필요한 설정요소 ### 
data = openpyxl.load_workbook(filepath+'data.xlsx') #로 데이터 시트 오픈(변경가능)
data_sheet_name = 'building' ## 데이터 파일 시트 
start_column = 'A' ## 데이터 파일 시트에서 값을 가져올 시작 열
end_column = 'AM' ## 데이터 파일 시트에서 값을 가져올 끝 열
iter = 4 ## 문서 생성을 할 데이터의 갯수 
data_sheet = data.get_sheet_by_name(data_sheet_name) #로 데이터 시트 객체 파일(land,시트명 : land)

### form sheet 에서 필요한 설정요소
form = openpyxl.load_workbook(filepath+'RPA_form.xlsx') #폼 데이터 시트 경로 지정(변경가능)
form_sheet_list = ['building_1','building_2'] ## 접근할 시트(추가 및 삭제 가능)
form_sheet_page_1 = form.get_sheet_by_name(form_sheet_list[0]) #폼 시트 페이지 1(building, 시트명 : land_1)
form_sheet_page_2 = form.get_sheet_by_name(form_sheet_list[1]) #폼 시트 페이지 2(building, 시트명 : land_2)

## 엑셀 PDF 연결
excel = win32com.client.Dispatch('Excel.Application')
excel.Visible = False


# 로 데이터 개별 객체를 데이터 시트에서 추출
for i in range(iter):
    sheet_start_cell = start_column + str(i+iter) 
    sheet_end_cell = end_column + str(i+iter)
    
    building_object = data_sheet[sheet_start_cell:sheet_end_cell] # 반복시 변경해야 됩니다
    data_list = []
    for row in building_object:
        for cell in row:
            data_list.append(cell.value)
    key_value = str(data_list[0])
    print(data_list)
    
    # 추출한 데이터를 양식에 매핑시켜주는 부분
    form_sheet_page_1['B2'] = data_list[0] ##관리번호
    form_sheet_page_1['D6'] = data_list[1] ##고유번호
    form_sheet_page_1['J6'] = data_list[2] ## 재산번호
    form_sheet_page_1['D7'] = data_list[3] ## 소재지
    form_sheet_page_1['J7'] = data_list[4] ## 재산명칭
    form_sheet_page_1['D8'] = data_list[5] ## 공유자수
    form_sheet_page_1['G8'] = data_list[6] ## 공유지분
    form_sheet_page_1['J8'] = data_list[7] ## 관리관
    form_sheet_page_1['D9'] = data_list[8] ## 재산구분
    form_sheet_page_1['J9'] = data_list[9] ## 위임기관
    form_sheet_page_1['D10'] = data_list[10]## 대장면적
    form_sheet_page_1['G10'] = data_list[11]## 대장지목
    form_sheet_page_1['J10'] = data_list[12] ## 연면적
    form_sheet_page_1['D11'] = data_list[13] ## 관련지번
    form_sheet_page_1['J11'] = data_list[14]## 용도
    form_sheet_page_1['D12'] = data_list[15]##공시지가
    form_sheet_page_1['G12'] = data_list[16] ##구조/지붕
    form_sheet_page_1['J12'] = data_list[17] ##층수
    form_sheet_page_1['D13'] = data_list[18] ## 용도지역
    form_sheet_page_1['G13'] = data_list[19] ## 용도지구
    form_sheet_page_1['J13'] = data_list[20] ## 용도구역
    form_sheet_page_1['C14'] = data_list[21] ## 이용현황
    form_sheet_page_1['C15'] = data_list[22] ## 토지이용계획

    # (예정) for문 삽입부분 
    # 이미지 입력부분 : 리사이즈 , 이미지 파일명 : 관리번호-1( ),관리번호-2( ), 관리번호-3( ),관리번호-4( )
    for j in range(4):
        img_file_name = filepath + '\\image\\' +  key_value +'-' +str(j+1) + input_image_format # 첫번째 페이지 이미지 이름 지정
        print(img_file_name)
        img = openpyxl.drawing.image.Image(img_file_name)
        img.width=430 # 이미지 리사이징, 가로.픽셀 단위입니다.
        img.height=286 # 이미지 리사이징 세로.픽셀 단위입니다.
        if(j==0):
            form_sheet_page_1.add_image(img,'B20')
        elif(j==1):
            form_sheet_page_1.add_image(img,'H20')
        elif(j==2):
            form_sheet_page_1.add_image(img,'B34')
        elif(j==3):
            form_sheet_page_1.add_image(img,'H34')

    #추출한 데이터를 매핑시켜주는 부분 :  building_2 sheet
    form_sheet_page_2['C4'] = data_list[23] ## 건축허가일
    form_sheet_page_2['G4'] = data_list[24] ## 사용승인일
    form_sheet_page_2['K4'] = data_list[25] ## 조사일
    form_sheet_page_2['C5'] = data_list[26] ## 연면적
    form_sheet_page_2['G5'] = data_list[27] ## 대지면적
    form_sheet_page_2['K5'] = data_list[28] ## 건축면적
    
    form_sheet_page_2['C8'] = data_list[29] ## 건폐율 - 건축물대장
    form_sheet_page_2['G8'] = data_list[30] ## 건폐율 -조사일기준
    form_sheet_page_2['K8'] = data_list[31] ## 건폐율 - 대지활용도
    form_sheet_page_2['C9'] = data_list[32] ## 용적율 - 건축물대장
    form_sheet_page_2['G9'] = data_list[33] ## 용적율 - 조사일 기준
    form_sheet_page_2['K9'] = data_list[34] ## 용적율 - 대지활용도

    form_sheet_page_2['C20'] = data_list[35] ## 토지표시사항
    form_sheet_page_2['G20'] = data_list[36] ## 불일치사항
    form_sheet_page_2['J20'] = data_list[37] ## 의견

    #output file 저장 부분
    output_file_name = filepath + '\\output\\' + saved_name + key_value +'.xlsx' ## data_list[0] 관리번호
    form.save(output_file_name)

    #output file PDF 저장
    wb = excel.WorkBooks.Open(output_file_name)
    ws_chart = wb.WorkSheets(form_sheet_list)
    ws_chart.Select()
    pdf_save_path = filepath + "\\output\\" + key_value + '.pdf'
    wb.ActiveSheet.ExportAsFixedFormat(0,pdf_save_path)
    wb.Close(True)

form.close()
data.close()
excel.Quit()